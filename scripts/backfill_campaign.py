"""Backfill de la campaña 'Beta DJs Junio 2026' a las tablas del dashboard.

Crea la fila en email_campaigns y carga los 861 envíos (desde los logs) con
su campaign_id. Para los 70 de hoy (AHORA) consulta el estado real en Resend.
Las tandas futuras quedan 'scheduled' y sus eventos entrarán por el webhook.

Idempotente: si la campaña ya existe, borra sus filas y recarga.
Uso: python3 scripts/backfill_campaign.py
"""
import csv
import json
import re
import time
import urllib.error
import urllib.request
from openpyxl import load_workbook

SLUG = "beta-djs-2026-06"
NAME = "Beta DJs — Junio 2026"
UA = "Mozilla/5.0 (Macintosh) Chrome/124"


def env(key):
    for l in open(".env.local", encoding="utf-8"):
        m = re.match(rf"\s*{key}\s*=\s*(.*)", l)
        if m:
            return m.group(1).strip().strip('"').strip("'")
    raise SystemExit(f"Falta {key} en .env.local")


SB_URL = env("NEXT_PUBLIC_SUPABASE_URL").rstrip("/")
SB_KEY = env("SUPABASE_SERVICE_ROLE_KEY")
RESEND_KEY = env("RESEND_API_KEY")


def sb(method, path, body=None, prefer=None):
    headers = {
        "apikey": SB_KEY,
        "Authorization": f"Bearer {SB_KEY}",
        "Content-Type": "application/json",
        "User-Agent": UA,
    }
    if prefer:
        headers["Prefer"] = prefer
    req = urllib.request.Request(
        f"{SB_URL}/rest/v1/{path}",
        data=json.dumps(body).encode() if body is not None else None,
        headers=headers,
        method=method,
    )
    with urllib.request.urlopen(req, timeout=40) as r:
        txt = r.read().decode()
        return json.loads(txt) if txt.strip() else None


def resend_status(eid):
    for i in range(4):
        try:
            req = urllib.request.Request(
                f"https://api.resend.com/emails/{eid}",
                headers={"Authorization": f"Bearer {RESEND_KEY}", "User-Agent": UA},
            )
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.load(r).get("last_event", "sent")
        except urllib.error.HTTPError as e:
            if e.code == 429 and i < 3:
                time.sleep(1.5 * (i + 1))
                continue
            return "sent"
        except Exception:
            return "sent"


# ── 1) Campaña (crear o reusar) ──────────────────────────────────────
existing = sb("GET", f"email_campaigns?slug=eq.{SLUG}&select=id")
if existing:
    cid = existing[0]["id"]
    print(f"Campaña ya existía ({cid}); limpio sus filas y recargo.")
    sb("DELETE", f"email_events?campaign_id=eq.{cid}", prefer="return=minimal")
    sb("DELETE", f"email_sends?campaign_id=eq.{cid}", prefer="return=minimal")
else:
    row = sb(
        "POST",
        "email_campaigns",
        {"slug": SLUG, "name": NAME, "status": "active", "total_recipients": 0},
        prefer="return=representation",
    )
    cid = row[0]["id"]
    print(f"Campaña creada: {cid}")

# ── 2) Leer logs + nombres ───────────────────────────────────────────
name_by = {}
wb = load_workbook("DROP_contactos_beta_cruce.xlsx", data_only=True)
ws = wb["Contactos"]
for r in range(2, ws.max_row + 1):
    nm, co = ws.cell(r, 2).value, ws.cell(r, 3).value
    if nm and co:
        name_by[str(co).strip().lower()] = str(nm).strip()

sends = {}  # resend_id -> dict (dedup por si aparece en ambos logs)
for path in ["scripts/envios_log.csv", "scripts/envios_log_retry.csv"]:
    try:
        for row in csv.DictReader(open(path, encoding="utf-8")):
            rid = row["id_o_error"]
            if len(rid) == 36 and "-" in rid:
                sched = row["scheduled_at"]
                sends[rid] = {
                    "correo": row["correo"],
                    "scheduled_at": None if sched == "AHORA" else sched,
                    "ahora": sched == "AHORA",
                }
    except FileNotFoundError:
        pass

print(f"Envíos OK en logs: {len(sends)}")

# ── 3) Estado real de los de hoy (AHORA) vía Resend ──────────────────
ahora_ids = [rid for rid, s in sends.items() if s["ahora"]]
print(f"Consultando estado de {len(ahora_ids)} envíos de hoy en Resend...")
status_by = {}
for rid in ahora_ids:
    status_by[rid] = resend_status(rid)
    time.sleep(0.2)

# ── 4) Construir filas e insertar (upsert por resend_id) ─────────────
rows_sends, rows_events = [], []
for rid, s in sends.items():
    last = status_by.get(rid, "scheduled" if s["scheduled_at"] else "queued")
    rows_sends.append({
        "campaign_id": cid,
        "resend_id": rid,
        "to_email": s["correo"],
        "to_name": name_by.get(s["correo"].lower()),
        "scheduled_at": s["scheduled_at"],
        "last_event": last,
    })
    if rid in status_by:  # historial p/ el feed de los de hoy
        rows_events.append({
            "resend_id": rid,
            "campaign_id": cid,
            "event_type": status_by[rid],
        })


def insert_batches(table, rows, prefer):
    for i in range(0, len(rows), 200):
        sb("POST", table, rows[i:i + 200], prefer=prefer)


insert_batches("email_sends", rows_sends, "resolution=merge-duplicates,return=minimal")
insert_batches("email_events", rows_events, "return=minimal")

# total_recipients
sb("PATCH", f"email_campaigns?id=eq.{cid}",
   {"total_recipients": len(rows_sends)}, prefer="return=minimal")

# ── 5) Resumen ───────────────────────────────────────────────────────
from collections import Counter
c = Counter(r["last_event"] for r in rows_sends)
print(f"\n✓ Backfill listo · campaña {SLUG} ({cid})")
print(f"  email_sends: {len(rows_sends)} | email_events (hoy): {len(rows_events)}")
print("  Estado de envíos:")
for k, v in c.most_common():
    print(f"    {k:12} {v}")
